#!/usr/bin/env python3
"""
Debian Package Dependency Analyzer

This script analyzes package dependencies from Debian trixie Sources.xz file.
"""

import os
import shutil
import gzip
import lzma
import urllib.request
import re
import pandas as pd
from datetime import datetime
from typing import Dict, List, Tuple, Optional
from collections import defaultdict, deque


class PackageDependencyAnalyzer:
    def __init__(self):
        self.result_dir = "result"
        self.sources_file = os.path.join(self.result_dir, "Sources")
        self.sources_url = "https://mirrors.ustc.edu.cn/debian/dists/trixie/main/source/Sources.xz"
    
    def init(self):
        """删除并重建result目录，下载最新的debian-trixie的Source.xz"""
        print("INFO: Initializing environment...")
        
        # 删除并重建result目录
        if os.path.exists(self.result_dir):
            shutil.rmtree(self.result_dir)
        os.makedirs(self.result_dir)
        
        # 下载Sources.xz文件
        sources_xz_path = os.path.join(self.result_dir, "Sources.xz")
        print(f"INFO: Downloading {self.sources_url}...")
        urllib.request.urlretrieve(self.sources_url, sources_xz_path)
        
        # 解压Sources.xz文件
        print("INFO: Extracting Sources.xz...")
        with lzma.open(sources_xz_path, 'rb') as f_in:
            with open(self.sources_file, 'wb') as f_out:
                shutil.copyfileobj(f_in, f_out)
        
        # 删除压缩文件
        os.remove(sources_xz_path)
        print("INFO: Environment initialization completed.")
    
    def _parse_sources_file(self) -> Dict[str, Dict[str, str]]:
        """解析Sources文件，返回包信息字典"""
        packages = {}
        current_package = {}
        current_package_name = None
        
        with open(self.sources_file, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                
                if not line:  # 空行，表示一个包的结束
                    if current_package_name and current_package:
                        packages[current_package_name] = current_package.copy()
                    current_package = {}
                    current_package_name = None
                    continue
                
                if line.startswith(' '):  # 继续字段
                    if 'current_field' in locals():
                        current_package[current_field] += ' ' + line.strip()
                    continue
                
                if ':' in line:
                    field, value = line.split(':', 1)
                    field = field.strip()
                    value = value.strip()
                    current_field = field
                    current_package[field] = value
                    
                    if field == 'Package':
                        current_package_name = value
        
        # 处理最后一个包
        if current_package_name and current_package:
            packages[current_package_name] = current_package.copy()
        
        return packages
    
    def _parse_dependencies(self, dep_string: str) -> List[str]:
        """解析依赖字符串，返回包名列表"""
        if not dep_string:
            return []
        
        # 移除版本约束和选择性依赖
        dep_string = re.sub(r'\([^)]*\)', '', dep_string)  # 移除版本约束
        dep_string = re.sub(r'\[[^\]]*\]', '', dep_string)  # 移除架构约束
        
        # 分割依赖项
        deps = []
        for dep in dep_string.split(','):
            # 处理OR依赖（|分隔），只取第一个
            dep = dep.split('|')[0].strip()
            if dep and not dep.startswith('$'):  # 排除变量
                deps.append(dep)
        
        return deps
    
    def analysis(self, target: str, no_all: str = "yes") -> Dict[str, List[str]]:
        """分析依赖于指定target的软件包"""
        print(f"INFO: 分析依赖于 '{target}' 的软件包...")
        
        if no_all.lower() == "yes":
            print("INFO: 过滤纯all架构的软件包")
        
        packages = self._parse_sources_file()
        result_dict = {}
        
        for pkg_name, pkg_info in packages.items():
            # 检查Build-Depends字段
            build_depends = pkg_info.get('Build-Depends', '')
            build_depends_indep = pkg_info.get('Build-Depends-Indep', '')
            
            all_deps = self._parse_dependencies(build_depends) + self._parse_dependencies(build_depends_indep)
            
            if target in all_deps:
                # 过滤纯all包
                architecture = pkg_info.get('Architecture', '')
                if no_all.lower() == "yes" and architecture == 'all':
                    continue
                
                # 提取包信息
                name = pkg_name
                section = pkg_info.get('Section', 'unknown')
                archs = architecture
                homepage = pkg_info.get('Homepage', '')
                
                result_dict[pkg_name] = [name, section, archs, homepage]
        
        print(f"INFO: 有{len(result_dict)}个软件包依赖于{target}")
        return result_dict
    
    def package_info(self, package_name: str) -> Dict[str, str]:
        """根据包名返回包的详细信息"""
        packages = self._parse_sources_file()
        
        if package_name not in packages:
            return {
                'category': 'unknown',
                'arch': 'unknown', 
                'homepage': ''
            }
        
        pkg_info = packages[package_name]
        return {
            'category': pkg_info.get('Section', 'unknown'),
            'arch': pkg_info.get('Architecture', 'unknown'),
            'homepage': pkg_info.get('Homepage', '')
        }
    
    def get_binary_packages(self, source_package_name: str) -> List[str]:
        """获取源码包产生的所有二进制包名"""
        packages = self._parse_sources_file()
        
        if source_package_name not in packages:
            return []
        
        pkg_info = packages[source_package_name]
        binary_field = pkg_info.get('Binary', '')
        
        if not binary_field:
            return []
        
        # 解析Binary字段，包名之间用逗号分隔
        binary_packages = []
        for pkg in binary_field.split(','):
            pkg = pkg.strip()
            if pkg:
                binary_packages.append(pkg)
        
        return binary_packages
    
    def export_to_excel(self, comprehensive_result: Dict[str, Dict[str, str]], target_list: List[str]) -> str:
        """将comprehensive_result导出到Excel文件"""
        # 准备数据
        data = []
        for pkg_name, info in comprehensive_result.items():
            data.append({
                '包名': pkg_name,
                '分类': info['category'],
                '架构': info['arch'],
                '主页': info['homepage'],
                '依赖链条': info['dependency_chain']
            })
        
        # 创建DataFrame
        df = pd.DataFrame(data)
        
        # 生成文件名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        target_names = "_".join(target_list[:3])  # 最多取前3个目标包名
        if len(target_list) > 3:
            target_names += f"_and_{len(target_list)-3}_more"
        filename = f"dependency_analysis_{target_names}_{timestamp}.xlsx"
        filepath = os.path.join(self.result_dir, filename)
        
        # 导出到Excel
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='依赖分析结果', index=False)
            
            # 获取工作表
            worksheet = writer.sheets['依赖分析结果']
            
            # 调整列宽
            worksheet.column_dimensions['A'].width = 25  # 包名
            worksheet.column_dimensions['B'].width = 20  # 分类
            worksheet.column_dimensions['C'].width = 15  # 架构
            worksheet.column_dimensions['D'].width = 50  # 主页
            worksheet.column_dimensions['E'].width = 80  # 依赖链条
            
            # 设置表头样式
            from openpyxl.styles import Font, PatternFill, Alignment
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 设置数据行样式
            for row in worksheet.iter_rows(min_row=2, max_row=len(data)+1):
                for cell in row:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
        
        return filepath
    
    def get_source_package_from_binary(self, binary_package: str) -> str:
        """根据二进制包名找到对应的源码包名"""
        packages = self._parse_sources_file()
        
        for source_pkg, pkg_info in packages.items():
            binary_field = pkg_info.get('Binary', '')
            if binary_field:
                binary_packages = [pkg.strip() for pkg in binary_field.split(',')]
                if binary_package in binary_packages:
                    return source_pkg
        
        # 如果没找到，假设源码包名和二进制包名相同
        return binary_package
    
    def _find_source_dependencies(self, target_binary: str, no_all: str, visited_sources: set = None, depth: int = 0, max_depth: int = 10) -> Tuple[Dict[str, List[str]], Dict[str, List[str]]]:
        """递归查找所有依赖的源码包（改进版）
        
        Args:
            target_binary: 目标二进制包名
            no_all: 是否过滤all架构包
            visited_sources: 已访问的源码包集合
            depth: 当前深度
            max_depth: 最大深度
            
        Returns:
            Tuple[source_deps_info, source_dependency_chains]: (源码包信息, 源码包依赖链条)
        """
        if visited_sources is None:
            visited_sources = set()
        
        if depth > max_depth:
            return {}, {}
        
        dependency_chain = {}
        all_source_deps = {}
        
        # 查找直接依赖于 target_binary 的所有源码包
        direct_source_deps = self.analysis(target_binary, no_all)
        
        if depth <= 2:
            print(f"INFO: 查找依赖于 '{target_binary}' 的源码包: 找到 {len(direct_source_deps)} 个")
        
        # 处理每个直接依赖的源码包
        for source_pkg in direct_source_deps.keys():
            if source_pkg in visited_sources:
                continue
                
            visited_sources.add(source_pkg)
            all_source_deps[source_pkg] = direct_source_deps[source_pkg]
            dependency_chain[source_pkg] = [source_pkg]  # 初始为源码包本身
            
            # 获取这个源码包产生的所有二进制包
            binary_packages = self.get_binary_packages(source_pkg)
            
            if depth <= 2:
                print(f"INFO: 源码包 '{source_pkg}' 产生 {len(binary_packages)} 个二进制包")
            
            # 对该源码包的所有二进制包进行递归分析
            for binary_pkg in binary_packages:
                if binary_pkg != target_binary:  # 避免循环
                    indirect_source_deps, indirect_chains = self._find_source_dependencies(
                        binary_pkg, no_all, visited_sources.copy(), depth + 1, max_depth
                    )
                    
                    # 合并间接依赖的源码包
                    for indirect_source, indirect_info in indirect_source_deps.items():
                        if indirect_source not in all_source_deps:
                            all_source_deps[indirect_source] = indirect_info
                            
                            # 构建依赖链条：当前源码包 -> 间接源码包
                            if indirect_source in indirect_chains:
                                dependency_chain[indirect_source] = [source_pkg] + indirect_chains[indirect_source]
                            else:
                                dependency_chain[indirect_source] = [source_pkg, indirect_source]
        
        return all_source_deps, dependency_chain
    
    def analyze_source_package_impact(self, source_package: str, no_all: str = "yes") -> Dict[str, Dict[str, any]]:
        """分析源码包构建成功后的影响（源码包模式）
        
        Args:
            source_package: 源码包名
            no_all: 是否过滤all架构包
            
        Returns:
            Dict: 受影响的源码包信息，包含依赖链条
        """
        print(f"\n分析源码包: {source_package}")
        print("=" * 50)
        
        # 获取该源码包产生的所有二进制包
        binary_packages = self.get_binary_packages(source_package)
        
        if not binary_packages:
            print(f"WARNING: 源码包 '{source_package}' 不存在或没有二进制包")
            return {}
        
        print(f"INFO: 源码包 '{source_package}' 产生 {len(binary_packages)} 个二进制包: {', '.join(binary_packages)}")
        
        # 对每个二进制包进行依赖分析
        all_affected_sources = {}
        
        for binary_pkg in binary_packages:
            print(f"\n分析二进制包: {binary_pkg}")
            print("-" * 30)
            
            # 查找依赖于该二进制包的源码包
            source_deps, source_chains = self._find_source_dependencies(binary_pkg, no_all, max_depth=5)
            
            print(f"找到 {len(source_deps)} 个依赖的源码包")
            
            # 处理每个受影响的源码包
            for affected_source, chain in source_chains.items():
                # 构建以源码包开头的依赖链条
                full_chain = f"{source_package} -> {' -> '.join(chain)}"
                
                if affected_source in all_affected_sources:
                    # 如果已存在，添加新的依赖链条
                    if full_chain not in all_affected_sources[affected_source]['chains']:
                        all_affected_sources[affected_source]['chains'].append(full_chain)
                else:
                    # 新的受影响源码包
                    pkg_info = self.package_info(affected_source)
                    all_affected_sources[affected_source] = {
                        'info': source_deps[affected_source],
                        'category': pkg_info['category'],
                        'arch': pkg_info['arch'],
                        'homepage': pkg_info['homepage'],
                        'chains': [full_chain]
                    }
                
                print(f"  {affected_source}: {source_package} -> {' -> '.join(chain)}")
        
        return all_affected_sources
    
    def main(self):
        """主流程控制（支持二进制包和源码包模式）"""
        # 第一步：初始化环境
        self.init()
        
        # 第二步：选择分析模式
        print("请选择分析模式:")
        print("1. 二进制包模式 - 以具体的二进制包为目标")
        print("2. 源码包模式 - 以源码包为目标，分析整个源码包的影响")
        
        mode_input = input("请输入模式编号 (1 或 2, 默认1): ").strip()
        if mode_input == "2":
            return self._main_source_mode()
        else:
            return self._main_binary_mode()
    
    def _main_binary_mode(self):
        """二进制包模式的主流程"""
        print("\n=== 二进制包模式 ===")
        
        # 接收目标包列表
        target_input = input("请输入要分析的目标二进制包名（用逗号分隔）: ").strip()
        target_list = [t.strip() for t in target_input.split(',') if t.strip()]
        
        if not target_list:
            print("ERROR: 未提供有效的目标包名")
            return
        
        # 接收过滤选项
        filter_input = input("是否过滤纯all包？(yes/no，默认yes): ").strip()
        if not filter_input:
            filter_input = "yes"
        
        # 分析每个目标包（源码包级别）
        final_source_deps = {}
        
        for target in target_list:
            print(f"\n分析目标二进制包: {target}")
            print("=" * 50)
            
            # 查承所有依赖的源码包
            source_deps, source_chains = self._find_source_dependencies(target, filter_input)
            
            print(f"找到 {len(source_deps)} 个依赖的源码包")
            
            # 显示结果
            for source_pkg, chain in source_chains.items():
                print(f"  {source_pkg}: {' -> '.join(chain)}")
                
                # 合并到最终结果中
                if source_pkg in final_source_deps:
                    # 如果已存在，添加新的依赖链条
                    chain_str = f"{target} -> {' -> '.join(chain)}"
                    if chain_str not in final_source_deps[source_pkg]['chains']:
                        final_source_deps[source_pkg]['chains'].append(chain_str)
                else:
                    # 新源码包
                    pkg_info = self.package_info(source_pkg)
                    final_source_deps[source_pkg] = {
                        'info': source_deps[source_pkg],
                        'category': pkg_info['category'],
                        'arch': pkg_info['arch'],
                        'homepage': pkg_info['homepage'],
                        'chains': [f"{target} -> {' -> '.join(chain)}"]
                    }
        
        return self._output_results(final_source_deps, target_list, "二进制包")
    
    def _main_source_mode(self):
        """源码包模式的主流程"""
        print("\n=== 源码包模式 ===")
        
        # 接收目标源码包列表
        target_input = input("请输入要分析的目标源码包名（用逗号分隔）: ").strip()
        target_list = [t.strip() for t in target_input.split(',') if t.strip()]
        
        if not target_list:
            print("ERROR: 未提供有效的目标源码包名")
            return
        
        # 接收过滤选项
        filter_input = input("是否过滤纯all包？(yes/no，默认yes): ").strip()
        if not filter_input:
            filter_input = "yes"
        
        # 分析每个目标源码包
        final_source_deps = {}
        
        for source_package in target_list:
            affected_sources = self.analyze_source_package_impact(source_package, filter_input)
            
            # 合并结果
            for source_pkg, data in affected_sources.items():
                if source_pkg in final_source_deps:
                    # 合并依赖链条
                    for chain in data['chains']:
                        if chain not in final_source_deps[source_pkg]['chains']:
                            final_source_deps[source_pkg]['chains'].append(chain)
                else:
                    final_source_deps[source_pkg] = data
        
        return self._output_results(final_source_deps, target_list, "源码包")
    
    def _output_results(self, final_source_deps: dict, target_list: list, mode_name: str):
        """输出最终结果"""
        print(f"\n分析完成！")
        print(f"共分析了 {len(target_list)} 个目标{mode_name}")
        print(f"找到 {len(final_source_deps)} 个唯一的受影响源码包")
        print(f"\n源码包依赖结果 ({mode_name}模式):")
        print("=" * 80)
        
        # 准备导出数据
        comprehensive_result = {}
        for source_pkg, data in final_source_deps.items():
            print(f"源码包: {source_pkg}")
            print(f"  分类: {data['category']}")
            print(f"  架构: {data['arch']}")
            print(f"  主页: {data['homepage']}")
            print(f"  依赖链条: {'; '.join(data['chains'])}")
            print("-" * 40)
            
            comprehensive_result[source_pkg] = {
                'category': data['category'],
                'arch': data['arch'],
                'homepage': data['homepage'],
                'dependency_chain': '; '.join(data['chains'])
            }
        
        # 导出到Excel
        if comprehensive_result:
            print(f"\n正在导出结果到Excel...")
            excel_filepath = self.export_to_excel(comprehensive_result, target_list)
            print(f"Excel文件已保存到: {excel_filepath}")
        
        return comprehensive_result


def main():
    """程序入口点"""
    analyzer = PackageDependencyAnalyzer()
    try:
        analyzer.main()
    except KeyboardInterrupt:
        print("\n\nINFO: 用户中断操作")
    except Exception as e:
        print(f"ERROR: {e}")


if __name__ == "__main__":
    main()